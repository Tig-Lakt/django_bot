from openpyxl import load_workbook
from config import XLSX_NAME_FILE


async def save_to_excel(data_list):
    """
    Открывает Excel файл и дописывает список из 5 элементов в новую строку,
    каждый элемент в свой столбец.
    """
    if len(data_list) != 5:
        raise ValueError("Список должен содержать ровно 5 элементов")

    wb = load_workbook(filename=XLSX_NAME_FILE)

    ws = wb.active

    next_row = ws.max_row + 1

    for col_index, value in enumerate(data_list, start=1):
        ws.cell(row=next_row, column=col_index, value=value)

    wb.save(XLSX_NAME_FILE)
    print(f"Данные успешно добавлены в строку {next_row}")